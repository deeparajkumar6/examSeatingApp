from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List
from datetime import datetime

class ExcelExportService:
    """Service for exporting seating arrangements to Excel format"""
    
    @staticmethod
    def generate_summary_excel(schedule_data: Dict) -> BytesIO:
        """Generate summary Excel with room-wise class distribution matching PDF format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Header styling
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16, color="FFFFFF")  # White text for red background
        room_font = Font(bold=True, size=12)
        class_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Red fill for title header (matching PDF #FF2222)
        red_fill = PatternFill(start_color="FF2222", end_color="FF2222", fill_type="solid")
        # Purple fill for room headers (matching PDF #B09FC6)
        purple_fill = PatternFill(start_color="B09FC6", end_color="B09FC6", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Title with red background
        title_cell = ws.cell(row=1, column=1, value=schedule_data.get('title', 'Exam Seating Arrangement'))
        title_cell.font = title_font
        title_cell.fill = red_fill
        title_cell.border = border
        ws.merge_cells('A1:C1')
        
        # Apply red background to merged cells
        for col in range(1, 4):
            ws.cell(row=1, column=col).fill = red_fill
            ws.cell(row=1, column=col).border = border
        
        # Date and session info
        ws['A2'] = f"Date: {schedule_data.get('date', '')}"
        ws['A3'] = f"Session: {schedule_data.get('session', '')}"
        
        # Process room data similar to PDF
        summary_data = ExcelExportService._process_summary_data(schedule_data.get('room_assignments', []))
        
        row = 5
        for room_index, room_data in enumerate(summary_data):
            # Room header with purple background
            room_cell = ws.cell(row=row, column=1, value=f"{room_index + 1}.Room No : {room_data['roomNumber']}({room_data['roomBuilding']})")
            room_cell.font = room_font
            room_cell.fill = purple_fill
            room_cell.border = border
            
            # Merge cells for room header
            ws.merge_cells(f'A{row}:C{row}')
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = purple_fill
                ws.cell(row=row, column=col).border = border
            
            row += 1
            
            # Class rows
            for class_info in room_data['classData']:
                # Class name
                class_cell = ws.cell(row=row, column=1, value=class_info['className'])
                class_cell.font = class_font
                class_cell.fill = white_fill
                class_cell.border = border
                
                # Roll range
                roll_cell = ws.cell(row=row, column=2, value=class_info['rollRange'])
                roll_cell.font = normal_font
                roll_cell.fill = white_fill
                roll_cell.border = border
                
                # Count
                count_cell = ws.cell(row=row, column=3, value=class_info['count'])
                count_cell.font = class_font
                count_cell.fill = white_fill
                count_cell.border = border
                count_cell.alignment = Alignment(horizontal='center')
                
                row += 1
            
            # Total row
            total_cell = ws.cell(row=row, column=1, value="Total")
            total_cell.font = class_font
            total_cell.fill = white_fill
            total_cell.border = border
            
            ws.cell(row=row, column=2, value="").fill = white_fill
            ws.cell(row=row, column=2).border = border
            
            total_count_cell = ws.cell(row=row, column=3, value=room_data['totalStudents'])
            total_count_cell.font = class_font
            total_count_cell.fill = white_fill
            total_count_cell.border = border
            total_count_cell.alignment = Alignment(horizontal='center')
            
            row += 2  # Add space between rooms
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def _process_summary_data(room_assignments: List) -> List:
        """Process room assignments into summary format matching PDF logic"""
        return [ExcelExportService._process_room_data(room) for room in room_assignments]
    
    @staticmethod
    def _process_room_data(room: Dict) -> Dict:
        """Process individual room data matching PDF logic"""
        # Group students by class
        class_summary = {}
        
        for student in room.get('students', []):
            class_name = student.get('className', 'Unknown')
            if class_name not in class_summary:
                class_summary[class_name] = {
                    'className': class_name,
                    'students': [],
                    'count': 0
                }
            
            roll_number = student.get('rollNumber', '')
            if roll_number:  # Only add non-empty roll numbers
                class_summary[class_name]['students'].append(roll_number)
            class_summary[class_name]['count'] += 1
        
        # Sort roll numbers and get range for each class
        class_data = []
        for class_info in class_summary.values():
            sorted_rolls = sorted([r for r in class_info['students'] if r])  # Filter out empty rolls
            
            if sorted_rolls:
                roll_range = f"{sorted_rolls[0]}-{sorted_rolls[-1]}" if len(sorted_rolls) > 1 else sorted_rolls[0]
            else:
                roll_range = 'Not assigned'
            
            class_data.append({
                'className': class_info['className'],
                'rollRange': roll_range,
                'count': class_info['count']
            })
        
        return {
            'roomNumber': room.get('room_number', ''),
            'roomBuilding': room.get('room_building', ''),
            'roomFloor': room.get('room_floor', ''),
            'totalCapacity': room.get('room_capacity', 0),
            'classData': class_data,
            'totalStudents': len(room.get('students', []))
        }
    
    @staticmethod
    def generate_detailed_excel(schedule_data: Dict) -> BytesIO:
        """Generate detailed Excel with one sheet per room"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for room in schedule_data.get('room_assignments', []):
            room_name = f"Room_{room.get('room_number', 'Unknown')}"
            ws = wb.create_sheet(title=room_name)
            
            # Header styling
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14, color="FFFFFF")  # White text for red background
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Red fill for title header (matching PDF #FF2222)
            red_fill = PatternFill(start_color="FF2222", end_color="FF2222", fill_type="solid")
            
            # Room title with red background
            title_cell = ws.cell(row=1, column=1, value=f"Room {room.get('room_number', '')} - {room.get('room_building', '')} Floor {room.get('room_floor', '')}")
            title_cell.font = title_font
            title_cell.fill = red_fill
            title_cell.border = border
            ws.merge_cells('A1:E1')
            
            # Apply red background to merged cells
            for col in range(1, 6):
                ws.cell(row=1, column=col).fill = red_fill
                ws.cell(row=1, column=col).border = border
            
            # Exam details
            ws['A2'] = f"Exam: {schedule_data.get('title', '')}"
            ws['A3'] = f"Date: {schedule_data.get('date', '')} | Session: {schedule_data.get('session', '')}"
            ws['A4'] = f"Capacity: {room.get('room_capacity', 0)} | Assigned: {len(room.get('students', []))}"
            
            # Student table headers
            headers = ['S.No', 'Roll Number', 'Student Name', 'Class', 'Language']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = header_font
                cell.border = border
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Student data
            students = room.get('students', [])
            for idx, student in enumerate(students, 1):
                row = 6 + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=student.get('rollNumber', 'Not assigned'))
                ws.cell(row=row, column=3, value=student.get('studentName', ''))
                ws.cell(row=row, column=4, value=student.get('className', ''))
                ws.cell(row=row, column=5, value=student.get('language', ''))
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
            
            # Auto-adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
