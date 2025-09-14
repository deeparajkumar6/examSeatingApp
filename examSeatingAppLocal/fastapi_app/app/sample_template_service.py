from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

class SampleTemplateService:
    """Service for generating sample Excel template for class data"""
    
    @staticmethod
    def generate_sample_template() -> BytesIO:
        """Generate sample Excel template with proper formatting and cell merging"""
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styling
        title_font = Font(bold=True, size=14, color='FFFFFF')
        header_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        red_fill = PatternFill(start_color='FF2222', end_color='FF2222', fill_type='solid')
        gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Sample data for demonstration
        sample_students = [
            [1, '', 'STUDENT NAME 1', 'I B.COM', 'I', 'TAMIL'],
            ['', '', '01/01/2005', '', '', ''],
            [2, '', 'STUDENT NAME 2', 'I B.COM', 'I', 'HINDI'],
            ['', '', '15/02/2005', '', '', ''],
            [3, '', 'STUDENT NAME 3', 'I B.COM', 'I', 'SANSKRIT'],
            ['', '', '20/03/2005', '', '', '']
        ]
        
        # Sheet 1: I Year Template with Language
        ws1 = wb.create_sheet('I_Year_Template')
        
        # Add academic year
        ws1.cell(row=4, column=2, value='I YEAR 2024-25').font = title_font
        ws1.cell(row=4, column=2).fill = red_fill
        ws1.merge_cells('B4:F4')
        for col in range(2, 7):
            ws1.cell(row=4, column=col).fill = red_fill
            ws1.cell(row=4, column=col).border = border
        
        # Headers
        headers = ['S.No', 'Register Number', 'Names with Date of Birth', 'Dept / Class', 'Shift', 'Language']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = gray_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Add sample data
        for i, student in enumerate(sample_students):
            row = 6 + i
            for col, value in enumerate(student, 1):
                ws1.cell(row=row, column=col, value=value)
                ws1.cell(row=row, column=col).border = border
        
        # Merge cells for student rows
        for i in range(0, len(sample_students), 2):
            row = 6 + i
            next_row = row + 1
            
            # Merge all columns except Names with DOB
            ws1.merge_cells(f'A{row}:A{next_row}')  # S.No
            ws1.merge_cells(f'B{row}:B{next_row}')  # Register Number
            ws1.merge_cells(f'D{row}:D{next_row}')  # Dept/Class
            ws1.merge_cells(f'E{row}:E{next_row}')  # Shift
            ws1.merge_cells(f'F{row}:F{next_row}')  # Language
        
        # Sheet 2: II/III Year Template without Language
        ws2 = wb.create_sheet('II_III_Year_Template')
        
        # Add academic year
        ws2.cell(row=4, column=2, value='II YEAR 2024-25').font = title_font
        ws2.cell(row=4, column=2).fill = red_fill
        ws2.merge_cells('B4:E4')
        for col in range(2, 6):
            ws2.cell(row=4, column=col).fill = red_fill
            ws2.cell(row=4, column=col).border = border
        
        # Headers (no language)
        headers = ['S.No', 'Register Number', 'Names with Date of Birth', 'Dept / Class', 'Shift']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = gray_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data for 2nd/3rd year (with roll numbers)
        sample_students_2nd = [
            [1, '22240001001', 'STUDENT NAME 1', 'II B.COM', 'I'],
            ['', '', '01/01/2003', '', ''],
            [2, '22240001002', 'STUDENT NAME 2', 'II B.COM', 'I'],
            ['', '', '15/02/2003', '', ''],
            [3, '22240001003', 'STUDENT NAME 3', 'II B.COM', 'I'],
            ['', '', '20/03/2003', '', '']
        ]
        
        # Add sample data
        for i, student in enumerate(sample_students_2nd):
            row = 6 + i
            for col, value in enumerate(student, 1):
                ws2.cell(row=row, column=col, value=value)
                ws2.cell(row=row, column=col).border = border
        
        # Merge cells for student rows
        for i in range(0, len(sample_students_2nd), 2):
            row = 6 + i
            next_row = row + 1
            
            # Merge all columns except Names with DOB
            ws2.merge_cells(f'A{row}:A{next_row}')  # S.No
            ws2.merge_cells(f'B{row}:B{next_row}')  # Register Number
            ws2.merge_cells(f'D{row}:D{next_row}')  # Dept/Class
            ws2.merge_cells(f'E{row}:E{next_row}')  # Shift
        
        # Auto-adjust column widths
        for ws in [ws1, ws2]:
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
